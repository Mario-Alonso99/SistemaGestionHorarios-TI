from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy

from apps.appDirection.administrators.models import Administrator
from apps.appDirection.administrators.forms import RegistroForm
from apps.appDirection.administrators.filters import AdministratorFilter

from openpyxl import Workbook
from django.http.response import HttpResponse, HttpResponseRedirect
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tablib import Dataset
from .resources import AdministratorResource
from .models import Administrator

# Create your views here.

class AdministratorCreate(CreateView):
    model = Administrator
    form_class = RegistroForm
    template_name = 'administrators/registrar.html'
    success_url = reverse_lazy('administrators:administrator_list')

class AdministratorList(ListView):
    queryset = Administrator.objects.filter(tipo='Administrator')
    queryset = Administrator.objects.order_by('estatus')
    template_name = 'administrators/administrator_list.html'
    paginate_by = 5

class AdministratorUpdate(UpdateView):
	model = Administrator
	form_class = RegistroForm
	template_name = 'administrators/registrar.html'
	success_url = reverse_lazy('administrators:administrator_list')

class AdministratorDelete(DeleteView):
	model = Administrator
	template_name = 'administrators/administrator_delete.html'
	success_url = reverse_lazy('administrators:administrator_list')

class AdministratorShow(DetailView):
    model = Administrator
    template_name = 'administrators/administrator_show.html'

def search(request):
    administrator_list = Administrator.objects.all()
    administrator_list = Administrator.objects.filter(tipo='Administrator')
    administrator_list = Administrator.objects.order_by('estatus')
    administrator_filter = AdministratorFilter(request.GET, queryset=administrator_list)
    return render(request, 'administrators/administrator_search.html', {'filter': administrator_filter})

class AdministratorReport(TemplateView):
    def get(self, request, *args, **kwargs):
        administrators = Administrator.objects.all()
        administrators= Administrator.objects.filter(tipo='Administrator')
        administrators = Administrator.objects.order_by('estatus')

        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name='Arial', size=16, bold= True)
        ws['B1'] = 'Reporte de administradores'
        
        nameHoja = 1
        ws.title = 'Hoja'+str(nameHoja)
        controlatorDatos = 4
        controlatorDatos +=1
        #Combinar Celdas
        ws.merge_cells('B1:H1')
        #Cambiar el Ancho de las Columnas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 30


        #Modificar el tamaño de las filas
        ws.row_dimensions[1].height = 25
        
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].font = Font(name='Arial', size=14, bold= True)
        ws['B3']= 'Matrícula'
        
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].font = Font(name='Arial', size=14, bold= True)
        ws['C3']= 'Nombre'
        
        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].font = Font(name='Arial', size=14, bold= True)
        ws['D3']= 'Apellidos'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['E3'].font = Font(name='Arial', size=14, bold= True)
        ws['E3']= 'Email'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['F3'].font = Font(name='Arial', size=14, bold= True)
        ws['F3']= 'Número de empleado'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['G3'].font = Font(name='Arial', size=14, bold= True)
        ws['G3']= 'Estatus'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['H3'].font = Font(name='Arial', size=14, bold= True)
        ws['H3']= 'Nombre de usuario'

        #Diseño de resultados en el reporte y consulta de datos
        cont = 4
        for administrator in administrators:
            ws.cell(row = cont, column = 2).value = administrator.matricula
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 2).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 3).value = administrator.first_name
            ws.cell(row = cont, column = 3).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 4).value = administrator.last_name
            ws.cell(row = cont, column = 4).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 5).value = administrator.email
            ws.cell(row = cont, column = 5).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 6).value = administrator.numero_empleado
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 6).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 7).value = administrator.estatus
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 7).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 8).value = administrator.username
            ws.cell(row = cont, column = 8).font = Font(name='Arial', size=10)
            
            cont +=1

        name_archivo = "ReporteAdministradores.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(name_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
"""
#Agregando codigo para la carga masiva
def importar(request):  
    if request.method == 'POST':  
        administrator_resource = AdministratorResource()  
        dataset = Dataset()  
        #print(dataset)  
        new_administrators = request.FILES['xlsfile']  
        #print(new_administrators)  
        imported_data = dataset.load(new_administrators.read())  
        #print(dataset)  
        result = administrator_resource.import_data(dataset, dry_run=True) # Test the data import  
        #print(result.has_errors())  
        if not result.has_errors():  
            administrator_resource.import_data(dataset, dry_run=False) # Actually import now
            return redirect(reverse_lazy('administrators:administrator_list'))
    return render(request, 'administrators/administrator_import.html')"""
