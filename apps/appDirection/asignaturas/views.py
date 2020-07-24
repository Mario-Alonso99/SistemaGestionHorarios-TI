from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy

from apps.appDirection.asignaturas.models import Asignatura
from apps.appDirection.asignaturas.forms import AsignaturaForm
from apps.appDirection.asignaturas.filters import AsignaturaFilter

from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Create your views here.
class AsignaturaCreate(CreateView):
    model = Asignatura
    form_class = AsignaturaForm
    template_name = 'asignaturas/asignatura_formCreate.html'
    success_url = reverse_lazy('asignaturas:asignatura_list')

class AsignaturaList(ListView):
    queryset = Asignatura.objects.order_by('especialidad', 'cuatrimestre')
    template_name = 'asignaturas/asignaturas_list.html'
    paginate_by = 5

class AsignaturaUpdate(UpdateView):
	model = Asignatura
	form_class = AsignaturaForm
	template_name = 'asignaturas/asignatura_formCreate.html'
	success_url = reverse_lazy('asignaturas:asignatura_list')

class AsignaturaDelete(DeleteView):
	model = Asignatura
	template_name = 'asignaturas/asignatura_delete.html'
	success_url = reverse_lazy('asignaturas:asignatura_list')

class AsignaturaShow(DetailView):
    model = Asignatura
    template_name = 'asignaturas/asignatura_show.html'

def search(request):
    asignatura_list = Asignatura.objects.all()
    asignatura_filter = AsignaturaFilter(request.GET, queryset=asignatura_list)
    return render(request, 'asignaturas/asignatura_search.html', {'filter': asignatura_filter})

class AsignaturaReport(TemplateView):
    def get(self, request, *args, **kwargs):
        asignaturas = Asignatura.objects.all()
        asignaturas = Asignatura.objects.order_by('especialidad', 'cuatrimestre')

        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name='Arial', size=16, bold= True)
        ws['B1'] = 'Reporte de Asignaturas'
        
        nameHoja = 1
        ws.title = 'Hoja'+str(nameHoja)
        controlatorDatos = 4
        controlatorDatos +=1
        #Combinar Celdas
        ws.merge_cells('B1:E1')
        #Cambiar el Ancho de las Columnas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
      

        #Modificar el tamaño de las filas
        ws.row_dimensions[1].height = 25
        
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].font = Font(name='Arial', size=14, bold= True)
        ws['B3']= 'Nombre'
        
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].font = Font(name='Arial', size=14, bold= True)
        ws['C3']= 'Especialidad'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].font = Font(name='Arial', size=14, bold= True)
        ws['D3']= 'Cuatrimestre'


        #Diseño de resultados en el reporte y consulta de datos
        cont = 4
        for asignatura in asignaturas:
            ws.cell(row = cont, column = 2).value = asignatura.nombre
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 2).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 3).value = asignatura.especialidad
            ws.cell(row = cont, column = 3).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 4).value = asignatura.cuatrimestre
            ws.cell(row = cont, column = 4).font = Font(name='Arial', size=10)


            cont +=1

        name_archivo = "Reporte de Asignaturas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(name_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response