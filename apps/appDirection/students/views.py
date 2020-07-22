from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy

from apps.appDirection.students.models import Student
from apps.appDirection.students.forms import StudentForm
from apps.appDirection.students.filters import StudentFilter

from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tablib import Dataset
from .resources import StudentResource
from .models import Student

#Lineas Extras
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from apps.appDirection.students.forms import RegistroForm

# Create your views here.
#Formulario de Registro de Usuario
class RegistroUsuario(CreateView):
    model = User
    form_class = RegistroForm
    template_name = 'students/registrar.html'
    success_url = reverse_lazy('students:student_list')

#Vista de Login
def login(request):
    return render(request, 'loginDirector/login.html')

class StudentCreate(CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/student_formCreate.html'
    success_url = reverse_lazy('students:student_list')

class StudentList(ListView):
    queryset = Student.objects.order_by('especialidad', 'cuatrimestre', 'grupo')
    template_name = 'students/students_list.html'
    paginate_by = 30

class StudentUpdate(UpdateView):
	model = Student
	form_class = StudentForm
	template_name = 'students/student_formCreate.html'
	success_url = reverse_lazy('students:student_list')

class StudentDelete(DeleteView):
	model = Student
	template_name = 'students/student_delete.html'
	success_url = reverse_lazy('students:student_list')

class StudentShow(DetailView):
    model = Student
    template_name = 'students/student_show.html'

def search(request):
    student_list = Student.objects.all()
    student_filter = StudentFilter(request.GET, queryset=student_list)
    return render(request, 'students/student_search.html', {'filter': student_filter})

class StudentReport(TemplateView):
    def get(self, request, *args, **kwargs):
        students = Student.objects.all()
        students = Student.objects.order_by('especialidad', 'cuatrimestre', 'grupo')

        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name='Arial', size=16, bold= True)
        ws['B1'] = 'Reporte de Estudiantes Inscritos'
        
        nameHoja = 1
        ws.title = 'Hoja'+str(nameHoja)
        controlatorDatos = 4
        controlatorDatos +=1
        #Combinar Celdas
        ws.merge_cells('B1:H1')
        #Cambiar el Ancho de las Columnas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 45
        ws.column_dimensions['H'].width = 20

        #Modificar el tamaño de las filas
        ws.row_dimensions[1].height = 25
        
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].font = Font(name='Arial', size=14, bold= True)
        ws['B3']= 'Matrícula'
        
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].font = Font(name='Arial', size=14, bold= True)
        ws['C3']= 'Nombre'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].font = Font(name='Arial', size=14, bold= True)
        ws['D3']= 'Especialidad'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['E3'].font = Font(name='Arial', size=14, bold= True)
        ws['E3']= 'Cuatrimestre'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['F3'].font = Font(name='Arial', size=14, bold= True)
        ws['F3'] = 'Grupo'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['G3'].font = Font(name='Arial', size=14, bold= True)
        ws['G3']= 'Email'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['H3'].font = Font(name='Arial', size=14, bold= True)
        ws['H3']= 'Estatus'

        #Diseño de resultados en el reporte y consulta de datos
        cont = 4
        for student in students:
            ws.cell(row = cont, column = 2).value = student.matricula
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 2).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 3).value = student.nombre
            ws.cell(row = cont, column = 3).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 4).value = student.especialidad
            ws.cell(row = cont, column = 4).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 5).value = student.cuatrimestre
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 5).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 6).value = student.grupo
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 6).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 7).value = student.email
            ws.cell(row = cont, column = 7).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 8).value = student.estatus
            ws.cell(row = cont, column = 8).font = Font(name='Arial', size=10)

            cont +=1

        name_archivo = "ReporteEstudiantesInscritos.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(name_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#Agregando codigo para la carga masiva
def importar(request):  
    if request.method == 'POST':  
        student_resource = StudentResource()  
        dataset = Dataset()  
        #print(dataset)  
        new_students = request.FILES['xlsfile']  
        #print(new_students)  
        imported_data = dataset.load(new_students.read())  
        #print(dataset)  
        result = student_resource.import_data(dataset, dry_run=True) # Test the data import  
        #print(result.has_errors())  
        if not result.has_errors():  
            student_resource.import_data(dataset, dry_run=False) # Actually import now
            return redirect(reverse_lazy('students:student_list'))
    return render(request, 'students/student_import.html')