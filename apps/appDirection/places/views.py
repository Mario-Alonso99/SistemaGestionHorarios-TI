from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy

from apps.appDirection.places.models import Place
from apps.appDirection.places.forms import PlaceForm
from apps.appDirection.places.filters import PlaceFilter

from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Create your views here.
class PlaceCreate(CreateView):
    model = Place
    form_class = PlaceForm
    template_name = 'places/place_formCreate.html'
    success_url = reverse_lazy('places:place_list')

class PlaceList(ListView):
    queryset = Place.objects.order_by('tipo', 'ubicacion', 'encargado')
    template_name = 'places/places_list.html'
    paginate_by = 5

class PlaceUpdate(UpdateView):
	model = Place
	form_class = PlaceForm
	template_name = 'places/place_formCreate.html'
	success_url = reverse_lazy('places:place_list')

class PlaceDelete(DeleteView):
	model = Place
	template_name = 'places/place_delete.html'
	success_url = reverse_lazy('places:place_list')

class PlaceShow(DetailView):
    model = Place
    template_name = 'places/place_show.html'

def search(request):
    place_list = Place.objects.all()
    place_filter = PlaceFilter(request.GET, queryset=place_list)
    return render(request, 'places/place_search.html', {'filter': place_filter})

class PlaceReport(TemplateView):
    def get(self, request, *args, **kwargs):
        places = Place.objects.all()
        places = Place.objects.order_by('tipo', 'ubicacion', 'encargado')

        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name='Arial', size=16, bold= True)
        ws['B1'] = 'Reporte de Aulas/Laboratorios'
        
        nameHoja = 1
        ws.title = 'Hoja'+str(nameHoja)
        controlatorDatos = 4
        controlatorDatos +=1
        #Combinar Celdas
        ws.merge_cells('B1:E1')
        #Cambiar el Ancho de las Columnas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 25

        #Modificar el tamaño de las filas
        ws.row_dimensions[1].height = 25
        
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].font = Font(name='Arial', size=14, bold= True)
        ws['B3']= 'Tipo'
        
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].font = Font(name='Arial', size=14, bold= True)
        ws['C3']= 'Encargado'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].font = Font(name='Arial', size=14, bold= True)
        ws['D3']= 'Ubicación'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['E3'].font = Font(name='Arial', size=14, bold= True)
        ws['E3']= 'Nombre'

        #Diseño de resultados en el reporte y consulta de datos
        cont = 4
        for place in places:
            ws.cell(row = cont, column = 2).value = place.tipo
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 2).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 3).value = place.encargado
            ws.cell(row = cont, column = 3).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 4).value = place.ubicacion
            ws.cell(row = cont, column = 4).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 5).value = place.nombre
            ws.cell(row = cont, column = 5).font = Font(name='Arial', size=10)

            cont +=1

        name_archivo = "Reporte de Aulas y Laboratorios.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(name_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response