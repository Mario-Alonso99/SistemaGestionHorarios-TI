from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy

from apps.appDirection.teachers.models import Teacher
from apps.appDirection.teachers.forms import RegistroForm
from apps.appDirection.teachers.filters import TeacherFilter

from openpyxl import Workbook
from django.http.response import HttpResponse, HttpResponseRedirect
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tablib import Dataset
from .resources import TeacherResource
from .models import Teacher

# Create your views here.

class TeacherCreate(CreateView):
    model = Teacher
    form_class = RegistroForm
    template_name = 'teachers/registrar.html'
    success_url = reverse_lazy('teachers:teacher_list')

class TeacherList(ListView):
    queryset = Teacher.objects.filter(tipo='Docente')
    queryset = Teacher.objects.order_by('grado_academico', 'tjornada')
    template_name = 'teachers/teachers_list.html'
    paginate_by = 8

class TeacherUpdate(UpdateView):
	model = Teacher
	form_class = RegistroForm
	template_name = 'teachers/registrar.html'
	success_url = reverse_lazy('teachers:teacher_list')

class TeacherDelete(DeleteView):
	model = Teacher
	template_name = 'teachers/teacher_delete.html'
	success_url = reverse_lazy('teachers:teacher_list')

class TeacherShow(DetailView):
    model = Teacher
    template_name = 'teachers/teacher_show.html'

def search(request):
    teacher_list = Teacher.objects.all()
    teacher_list = Teacher.objects.filter(tipo='Docente')
    teacher_list = Teacher.objects.order_by('grado_academico', 'tjornada')
    teacher_filter = TeacherFilter(request.GET, queryset=teacher_list)
    return render(request, 'teachers/teacher_search.html', {'filter': teacher_filter})

class TeacherReport(TemplateView):
    def get(self, request, *args, **kwargs):
        teachers = Teacher.objects.all()
        teachers= Teacher.objects.filter(tipo='Docente')
        teachers = Teacher.objects.order_by('grado_academico','tjornada', 'estatus')

        wb = Workbook()
        ws = wb.active
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name='Arial', size=16, bold= True)
        ws['B1'] = 'Reporte de docentes'
        
        nameHoja = 1
        ws.title = 'Hoja'+str(nameHoja)
        controlatorDatos = 4
        controlatorDatos +=1
        #Combinar Celdas
        ws.merge_cells('B1:J1')
        #Cambiar el Ancho de las Columnas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 35
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30



        #Modificar el tamaño de las filas
        ws.row_dimensions[1].height = 25
        
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['B3'].font = Font(name='Arial', size=14, bold= True)
        ws['B3']= 'Matrícula'
        
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['C3'].font = Font(name='Arial', size=14, bold= True)
        ws['C3']= 'Nombre'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['D3'].font = Font(name='Arial', size=14, bold= True)
        ws['D3']= 'Apellidos'
        
        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['E3'].font = Font(name='Arial', size=14, bold= True)
        ws['E3']= 'Email'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['F3'].font = Font(name='Arial', size=14, bold= True)
        ws['F3']= 'Grado académico'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['G3'].font = Font(name='Arial', size=14, bold= True)
        ws['G3']= 'Tipo'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['H3'].font = Font(name='Arial', size=14, bold= True)
        ws['H3']= 'Número de profesor'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['I3'].font = Font(name='Arial', size=14, bold= True)
        ws['I3']= 'Estatus'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        ws['J3'].font = Font(name='Arial', size=14, bold= True)
        ws['J3']= 'Nombre de usuario'




        #Diseño de resultados en el reporte y consulta de datos
        cont = 4
        for teacher in teachers:
            ws.cell(row = cont, column = 2).value = teacher.matricula
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 2).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 3).value = teacher.first_name
            ws.cell(row = cont, column = 3).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 4).value = teacher.last_name
            ws.cell(row = cont, column = 4).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 5).value = teacher.email
            ws.cell(row = cont, column = 5).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 6).value = teacher.grado_academico
            ws.cell(row = cont, column = 6).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 7).value = teacher.tjornada
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 7).font = Font(name='Arial', size=10)


            ws.cell(row = cont, column = 8).value = teacher.numero_empleado
            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = cont, column = 8).font = Font(name='Arial', size=10)
            

            ws.cell(row = cont, column = 9).value = teacher.estatus
            ws.cell(row = cont, column = 9).font = Font(name='Arial', size=10)

            ws.cell(row = cont, column = 10).value = teacher.username
            ws.cell(row = cont, column = 10).font = Font(name='Arial', size=10)

            cont +=1

        name_archivo = "ReporteDocentes.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(name_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#Agregando codigo para la carga masiva
def importar(request):  
    if request.method == 'POST':  
        teacher_resource = TeacherResource()  
        dataset = Dataset()  
        #print(dataset)  
        new_teachers = request.FILES['xlsfile']  
        #print(new_teachers)  
        imported_data = dataset.load(new_teachers.read())  
        #print(dataset)  
        result = teacher_resource.import_data(dataset, dry_run=True) # Test the data import  
        #print(result.has_errors())  
        if not result.has_errors():  
            teacher_resource.import_data(dataset, dry_run=False) # Actually import now
            return redirect(reverse_lazy('teachers:teacher_list'))
    return render(request, 'teachers/teacher_import.html')